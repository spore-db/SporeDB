"""Excel reader with multi-sheet classification and batch-per-sheet mode."""

from __future__ import annotations

import contextlib
import time
from datetime import datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

from sporedb.ingestion.column_mapper import detect_columns
from sporedb.ingestion.csv_reader import (
    _detect_unit_conversions,
    _parse_rows_to_records,
    _store_original_values,
)
from sporedb.ingestion.result import ColumnMapping, ImportResult
from sporedb.ingestion.timestamp import (
    TIMESTAMP_COLUMN_NAMES,
    detect_elapsed_unit,
    detect_timestamp_column,
    parse_timestamps,
)
from sporedb.models.batch import Batch
from sporedb.storage.batch_store import BatchStore
from sporedb.storage.engine import StorageEngine
from sporedb.storage.ts_store import TimeSeriesStore


class SheetType(StrEnum):
    """Classification of Excel sheet content."""

    METADATA = "metadata"
    TELEMETRY = "telemetry"
    ASSAY = "assay"
    UNKNOWN = "unknown"


# Sheet names that indicate metadata content
METADATA_SHEET_NAMES: set[str] = {
    "metadata",
    "info",
    "setup",
    "batch_info",
    "run_info",
    "batch_metadata",
    "experiment_info",
}

# Column name keywords that indicate assay data
ASSAY_INDICATORS: set[str] = {
    "method",
    "uncertainty",
    "hplc",
    "lc-ms",
    "cell_count",
    "assay",
}


def classify_sheet(sheet_name: str, headers: list[str]) -> SheetType:
    """Classify a sheet based on its name and column headers.

    Priority:
    1. Sheet name matches known metadata names -> METADATA
    2. Headers contain assay indicators + timestamp -> ASSAY
    3. Headers contain timestamp column -> TELEMETRY
    4. Otherwise -> UNKNOWN
    """
    # Check sheet name against metadata names (case-insensitive)
    if sheet_name.lower().strip() in METADATA_SHEET_NAMES:
        return SheetType.METADATA

    headers_lower = {h.lower().strip() for h in headers}

    # Check for assay indicators
    has_assay_indicator = bool(headers_lower & ASSAY_INDICATORS)
    has_timestamp = bool(headers_lower & TIMESTAMP_COLUMN_NAMES)

    if has_assay_indicator and has_timestamp:
        return SheetType.ASSAY

    # Check for timestamp column (telemetry)
    if has_timestamp:
        return SheetType.TELEMETRY

    # Also check if "assay" is in the sheet name
    if "assay" in sheet_name.lower():
        return SheetType.ASSAY

    return SheetType.UNKNOWN


def detect_batch_per_sheet_mode(workbook_sheets: dict[str, list[str]]) -> bool:
    """Detect if a workbook uses batch-per-sheet mode.

    Returns True if all sheets have identical column structures
    (same headers after case normalization).
    """
    if len(workbook_sheets) < 2:
        return False

    normalized_headers = [
        [h.lower().strip() for h in headers] for headers in workbook_sheets.values()
    ]

    first = normalized_headers[0]
    return all(h == first for h in normalized_headers[1:])


def _parse_metadata_sheet(rows: list[list[Any]]) -> dict[str, str]:
    """Parse key-value pairs from a metadata sheet (column A=key, column B=value)."""
    metadata: dict[str, str] = {}
    for row in rows:
        if len(row) >= 2 and row[0] is not None:
            key = str(row[0]).strip().lower()
            value = str(row[1]).strip() if row[1] is not None else ""
            if key:
                metadata[key] = value
    return metadata


def _read_excel_sheets(file_path: Path) -> dict[str, tuple[list[str], list[list[Any]]]]:
    """Read all sheets from an Excel file using openpyxl read_only mode.

    Returns:
        Dict of {sheet_name: (headers, data_rows)}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        result: dict[str, tuple[list[str], list[list[Any]]]] = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows()

            # Read all rows
            all_rows: list[list[Any]] = []
            for row in rows_iter:
                cell_values = []
                for cell in row:
                    val = cell.value
                    # Handle MergedCell (None values)
                    if val is None:
                        cell_values.append("")
                    else:
                        cell_values.append(str(val))
                all_rows.append(cell_values)

            if not all_rows:
                result[sheet_name] = ([], [])
                continue

            headers = all_rows[0]
            data_rows = all_rows[1:]
            result[sheet_name] = (headers, data_rows)

        return result
    finally:
        wb.close()


def _apply_metadata_to_batch(batch: Batch, metadata: dict[str, str]) -> Batch:
    """Apply parsed metadata to a Batch model."""
    meta = batch.metadata
    if "strain" in metadata:
        meta.strain = metadata["strain"]
    if "media" in metadata:
        meta.media = metadata["media"]
    if "scale_liters" in metadata:
        with contextlib.suppress(ValueError, TypeError):
            meta.scale_liters = float(metadata["scale_liters"])
    if "operator" in metadata:
        meta.operator = metadata["operator"]

    # Store remaining keys in extra
    known_keys = {"strain", "media", "scale_liters", "operator"}
    for key, value in metadata.items():
        if key not in known_keys:
            meta.extra[key] = value

    batch.metadata = meta
    return batch


def import_excel(
    file_path: Path | str,
    batch_name: str,
    engine: StorageEngine,
    mapping: ColumnMapping | None = None,
    inoculation_ts: datetime | None = None,
    custom_vocab: dict[str, list[str]] | None = None,
    batch_per_sheet: bool | None = None,
) -> ImportResult | list[ImportResult]:
    """Import an Excel file into SporeDB.

    Supports multi-sheet classification (metadata/telemetry/assay) and
    batch-per-sheet mode for workbooks with one run per sheet.

    Args:
        file_path: Path to the .xlsx file.
        batch_name: Base name for the batch(es).
        engine: Storage engine instance.
        mapping: Pre-built column mapping (two-pass flow).
        inoculation_ts: Reference timestamp for elapsed time conversion.
        custom_vocab: Additional vocabulary for column matching.
        batch_per_sheet: Force batch-per-sheet mode. None = auto-detect.

    Returns:
        Single ImportResult or list[ImportResult] if batch_per_sheet mode.

    Raises:
        ValueError: If file is too large, empty, or invalid.
    """
    start_time = time.time()

    # 1. Validate path (T-02-05)
    file_path = Path(file_path).resolve()
    if not file_path.is_file():
        raise ValueError(f"File not found or not a regular file: {file_path}")

    # Zip bomb protection (T-02-07): 100MB limit
    if file_path.stat().st_size > 100_000_000:
        raise ValueError(f"File exceeds 100MB size limit: {file_path}")

    # 2. Read all sheets
    sheets = _read_excel_sheets(file_path)
    if not sheets:
        raise ValueError(f"No sheets found in Excel file: {file_path}")

    # 3. Auto-detect batch_per_sheet if not specified
    if batch_per_sheet is None:
        sheet_headers = {name: headers for name, (headers, _) in sheets.items()}
        batch_per_sheet = detect_batch_per_sheet_mode(sheet_headers)

    # 4. Batch-per-sheet mode
    if batch_per_sheet:
        return _import_batch_per_sheet(
            sheets,
            batch_name,
            engine,
            mapping,
            inoculation_ts,
            custom_vocab,
            start_time,
        )

    # 5. Standard multi-sheet mode
    return _import_multi_sheet(
        sheets, batch_name, engine, mapping, inoculation_ts, custom_vocab, start_time
    )


def _import_batch_per_sheet(
    sheets: dict[str, tuple[list[str], list[list[Any]]]],
    batch_name: str,
    engine: StorageEngine,
    mapping: ColumnMapping | None,
    inoculation_ts: datetime | None,
    custom_vocab: dict[str, list[str]] | None,
    start_time: float,
) -> list[ImportResult]:
    """Import each sheet as a separate batch."""
    results: list[ImportResult] = []
    batch_store = BatchStore(engine)
    ts_store = TimeSeriesStore(engine)

    for sheet_name, (headers, data_rows) in sheets.items():
        if not data_rows:
            continue

        # Auto-detect mapping per sheet if not provided
        sheet_mapping = mapping
        if sheet_mapping is None:
            sheet_mapping = detect_columns(
                headers, data_rows[:5], custom_vocab=custom_vocab
            )

        # Detect timestamps
        ts_col, is_elapsed = detect_timestamp_column(headers, data_rows[:5])
        ts_col_idx = headers.index(ts_col) if ts_col in headers else None
        if ts_col_idx is None:
            continue

        ts_values = [row[ts_col_idx] for row in data_rows if ts_col_idx < len(row)]

        if is_elapsed and inoculation_ts is None:
            raise ValueError(
                f"Elapsed time in sheet '{sheet_name}' but no inoculation_ts provided."
            )

        elapsed_unit = detect_elapsed_unit(ts_col) if is_elapsed else "h"
        timestamps = parse_timestamps(
            ts_values,
            is_elapsed,
            reference_ts=inoculation_ts,
            elapsed_unit=elapsed_unit,
        )

        # Unit conversions
        unit_conversions = _detect_unit_conversions(headers, sheet_mapping, data_rows)

        # Create batch
        full_name = f"{batch_name}-{sheet_name}"
        batch = Batch(name=full_name)
        batch = batch_store.create_batch(batch)

        # Store original values
        _store_original_values(engine.data_root, batch.batch_id, headers, data_rows)

        # Convert and persist
        records, warnings, units_converted = _parse_rows_to_records(
            data_rows,
            headers,
            sheet_mapping,
            batch.batch_id,
            timestamps,
            unit_conversions,
        )
        if records:
            ts_store.append_telemetry(records)  # type: ignore[arg-type]

        elapsed = time.time() - start_time
        results.append(
            ImportResult(
                batch_id=batch.batch_id,
                rows_imported=len(data_rows),
                columns_mapped=sheet_mapping.variable_mappings,
                units_converted=units_converted,
                warnings=warnings,
                elapsed_seconds=elapsed,
            )
        )

    return results


def _import_multi_sheet(
    sheets: dict[str, tuple[list[str], list[list[Any]]]],
    batch_name: str,
    engine: StorageEngine,
    mapping: ColumnMapping | None,
    inoculation_ts: datetime | None,
    custom_vocab: dict[str, list[str]] | None,
    start_time: float,
) -> ImportResult:
    """Import a multi-sheet workbook as a single batch with classified sheets."""
    batch_store = BatchStore(engine)
    ts_store = TimeSeriesStore(engine)

    # Classify all sheets
    classified: dict[str, tuple[SheetType, list[str], list[list[Any]]]] = {}
    for sheet_name, (headers, data_rows) in sheets.items():
        sheet_type = classify_sheet(sheet_name, headers)
        classified[sheet_name] = (sheet_type, headers, data_rows)

    # Create batch first
    batch = Batch(name=batch_name)

    # Extract metadata from metadata sheets
    for _sheet_name, (sheet_type, headers, data_rows) in classified.items():
        if sheet_type == SheetType.METADATA:
            # Metadata sheets are key-value; include first row (treated as "headers")
            all_rows = [headers] + data_rows
            metadata = _parse_metadata_sheet(all_rows)
            batch = _apply_metadata_to_batch(batch, metadata)

    batch = batch_store.create_batch(batch)

    total_rows = 0
    all_warnings: list[str] = []
    all_columns_mapped: dict[str, str] = {}
    all_units_converted: dict[str, tuple[str, str]] = {}

    # Process telemetry and assay sheets
    for sheet_name, (sheet_type, headers, data_rows) in classified.items():
        if sheet_type in (SheetType.METADATA, SheetType.UNKNOWN):
            continue
        if not data_rows:
            continue

        # Determine data_type
        data_type = "assay" if sheet_type == SheetType.ASSAY else "telemetry"

        # Auto-detect mapping if not provided
        sheet_mapping = mapping
        if sheet_mapping is None:
            sheet_mapping = detect_columns(
                headers, data_rows[:5], custom_vocab=custom_vocab
            )

        # Detect timestamps
        try:
            ts_col, is_elapsed = detect_timestamp_column(headers, data_rows[:5])
        except ValueError:
            all_warnings.append(
                f"Sheet '{sheet_name}': no timestamp column found, skipping"
            )
            continue

        ts_col_idx = headers.index(ts_col) if ts_col in headers else None
        if ts_col_idx is None:
            continue

        ts_values = [row[ts_col_idx] for row in data_rows if ts_col_idx < len(row)]

        if is_elapsed and inoculation_ts is None:
            all_warnings.append(
                f"Sheet '{sheet_name}': elapsed time without inoculation_ts, skipping"
            )
            continue

        elapsed_unit = detect_elapsed_unit(ts_col) if is_elapsed else "h"
        timestamps = parse_timestamps(
            ts_values,
            is_elapsed,
            reference_ts=inoculation_ts,
            elapsed_unit=elapsed_unit,
        )

        # Unit conversions
        unit_conversions = _detect_unit_conversions(headers, sheet_mapping, data_rows)

        # Store original values
        _store_original_values(engine.data_root, batch.batch_id, headers, data_rows)

        # Convert and persist
        records, warnings, units_converted = _parse_rows_to_records(
            data_rows,
            headers,
            sheet_mapping,
            batch.batch_id,
            timestamps,
            unit_conversions,
            data_type,
        )

        if records:
            if data_type == "assay":
                ts_store.append_assay(records)  # type: ignore[arg-type]
            else:
                ts_store.append_telemetry(records)  # type: ignore[arg-type]

        total_rows += len(data_rows)
        all_warnings.extend(warnings)
        all_columns_mapped.update(sheet_mapping.variable_mappings)
        all_units_converted.update(units_converted)

    elapsed = time.time() - start_time
    return ImportResult(
        batch_id=batch.batch_id,
        rows_imported=total_rows,
        columns_mapped=all_columns_mapped,
        units_converted=all_units_converted,
        warnings=all_warnings,
        elapsed_seconds=elapsed,
    )

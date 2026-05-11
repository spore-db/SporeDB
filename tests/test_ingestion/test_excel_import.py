"""Tests for Excel reader with multi-sheet classification."""

from __future__ import annotations

from datetime import UTC, datetime

import openpyxl
import pytest

from sporedb.ingestion.excel_reader import (
    SheetType,
    classify_sheet,
    detect_batch_per_sheet_mode,
    import_excel,
)
from sporedb.ingestion.result import ColumnMapping, ImportResult
from sporedb.storage.engine import StorageEngine
from sporedb.storage.ts_store import TimeSeriesStore


@pytest.fixture
def engine(data_root):
    return StorageEngine(data_root)


@pytest.fixture
def sample_workbook_path(tmp_path):
    """Create a sample Excel workbook with metadata + telemetry + assay sheets."""
    wb = openpyxl.Workbook()

    # Metadata sheet
    ws_meta = wb.active
    ws_meta.title = "metadata"
    ws_meta.append(["strain", "CHO-K1"])
    ws_meta.append(["media", "CD-CHO"])
    ws_meta.append(["scale_liters", "5.0"])
    ws_meta.append(["operator", "Dr. Smith"])

    # Telemetry sheet
    ws_telem = wb.create_sheet("telemetry")
    ws_telem.append(["timestamp", "pH", "DO_%", "temp_C", "OD600"])
    ws_telem.append(["2026-04-20 08:00:00", "7.0", "95.2", "37.0", "0.5"])
    ws_telem.append(["2026-04-20 08:15:00", "6.95", "88.1", "37.1", "0.8"])
    ws_telem.append(["2026-04-20 08:30:00", "6.88", "82.0", "37.0", "1.2"])
    ws_telem.append(["2026-04-20 08:45:00", "6.82", "75.3", "36.9", "1.8"])
    ws_telem.append(["2026-04-20 09:00:00", "6.75", "68.5", "37.0", "2.5"])

    # Assay sheet
    ws_assay = wb.create_sheet("assay_results")
    ws_assay.append(["timestamp", "glucose", "lactate", "method"])
    ws_assay.append(["2026-04-20 08:00:00", "20.0", "0.5", "HPLC"])
    ws_assay.append(["2026-04-20 08:30:00", "16.0", "1.2", "HPLC"])
    ws_assay.append(["2026-04-20 09:00:00", "10.1", "2.0", "HPLC"])

    path = tmp_path / "sample_workbook.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def sample_batch_per_sheet_path(tmp_path):
    """Create workbook with identical-structure sheets (batch-per-sheet mode)."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Run-A"
    ws1.append(["time_h", "pH", "DO_%", "OD600"])
    ws1.append(["0.0", "7.0", "95.2", "0.5"])
    ws1.append(["0.25", "6.95", "88.1", "0.8"])

    ws2 = wb.create_sheet("Run-B")
    ws2.append(["time_h", "pH", "DO_%", "OD600"])
    ws2.append(["0.0", "7.1", "94.0", "0.4"])
    ws2.append(["0.25", "6.98", "87.5", "0.7"])

    ws3 = wb.create_sheet("Run-C")
    ws3.append(["time_h", "pH", "DO_%", "OD600"])
    ws3.append(["0.0", "6.9", "96.1", "0.6"])
    ws3.append(["0.25", "6.85", "89.0", "0.9"])

    path = tmp_path / "batch_per_sheet.xlsx"
    wb.save(path)
    return path


class TestSheetClassification:
    """Test sheet type classification."""

    def test_sheet_type_enum_values(self):
        assert SheetType.METADATA == "metadata"
        assert SheetType.TELEMETRY == "telemetry"
        assert SheetType.ASSAY == "assay"
        assert SheetType.UNKNOWN == "unknown"

    def test_classify_metadata_by_name(self):
        assert classify_sheet("metadata", ["strain", "value"]) == SheetType.METADATA

    def test_classify_setup_as_metadata(self):
        assert classify_sheet("setup", ["key", "value"]) == SheetType.METADATA

    def test_classify_telemetry_by_headers(self):
        result = classify_sheet("Sheet1", ["time_h", "pH", "DO"])
        assert result == SheetType.TELEMETRY

    def test_classify_assay_by_headers(self):
        result = classify_sheet(
            "assay_data", ["time", "glucose", "method", "uncertainty"]
        )
        assert result == SheetType.ASSAY

    def test_classify_unknown_fallback(self):
        result = classify_sheet("random_sheet", ["col_a", "col_b", "col_c"])
        assert result == SheetType.UNKNOWN


class TestBatchPerSheetDetection:
    """Test batch-per-sheet mode detection."""

    def test_identical_headers_returns_true(self):
        sheets = {
            "Run-A": ["time_h", "pH", "DO_%", "OD600"],
            "Run-B": ["time_h", "pH", "DO_%", "OD600"],
            "Run-C": ["time_h", "pH", "DO_%", "OD600"],
        }
        assert detect_batch_per_sheet_mode(sheets) is True

    def test_different_headers_returns_false(self):
        sheets = {
            "metadata": ["key", "value"],
            "telemetry": ["timestamp", "pH", "DO_%"],
            "assay": ["timestamp", "glucose", "method"],
        }
        assert detect_batch_per_sheet_mode(sheets) is False


class TestImportExcel:
    """Test full Excel import pipeline."""

    def test_import_multi_sheet_workbook(self, engine, sample_workbook_path):
        result = import_excel(
            sample_workbook_path,
            "CHO-Run-001",
            engine=engine,
        )
        assert isinstance(result, ImportResult)
        assert result.rows_imported > 0

    def test_metadata_extracted_from_sheet(self, engine, sample_workbook_path):
        result = import_excel(
            sample_workbook_path,
            "CHO-Run-002",
            engine=engine,
        )
        # Batch metadata should include strain from metadata sheet
        from sporedb.storage.batch_store import BatchStore

        batch_store = BatchStore(engine)
        batch = batch_store.get_batch(result.batch_id)
        assert batch is not None
        assert batch.metadata.strain == "CHO-K1"

    def test_telemetry_rows_imported(self, engine, sample_workbook_path):
        result = import_excel(
            sample_workbook_path,
            "CHO-Run-003",
            engine=engine,
        )
        ts_store = TimeSeriesStore(engine)
        df = ts_store.get_telemetry(result.batch_id)
        assert not df.empty


class TestImportExcelBatchPerSheet:
    """Test batch-per-sheet mode."""

    def test_batch_per_sheet_returns_list(self, engine, sample_batch_per_sheet_path):
        inoc_ts = datetime(2026, 4, 20, 8, 0, 0, tzinfo=UTC)
        results = import_excel(
            sample_batch_per_sheet_path,
            "Multi-Run",
            engine=engine,
            batch_per_sheet=True,
            inoculation_ts=inoc_ts,
        )
        assert isinstance(results, list)
        assert len(results) == 3
        for r in results:
            assert r.rows_imported > 0


class TestImportExcelTwoPass:
    """Test two-pass flow with provided ColumnMapping."""

    def test_provided_mapping_used(self, engine, sample_workbook_path):
        mapping = ColumnMapping(
            timestamp_col="timestamp",
            variable_mappings={
                "pH": "ph",
                "DO_%": "dissolved_oxygen",
            },
            unit_mappings={},
            unmapped_cols=["temp_C", "OD600"],
            confidence={"pH": 1.0, "DO_%": 1.0},
        )
        result = import_excel(
            sample_workbook_path,
            "CHO-TwoPass",
            engine=engine,
            mapping=mapping,
        )
        assert result.rows_imported > 0
        # Only pH and DO should be in columns_mapped
        assert len(result.columns_mapped) == 2
